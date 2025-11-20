import re
import csv
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl import Workbook

RE_HAS_LOTTO_COLON = re.compile(r"(?i)\blotto:")
RE_HAS_LOTTO_WORD  = re.compile(r"(?i)\blotto\b")  # matches '... Lotto ...'

_num_token = re.compile(r"[-(]?\s*[\d.,]+(?:\s*%)?\s*[)]?")

def to_float(v):
    """
    Convert any cell value to float.
    - Handles commas and dots as decimal separators
    - Ignores text, blanks, and non-numerics
    - If multiple numbers in a string -> takes only the FIRST one
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None

    # find first number-like token
    m = _num_token.search(v)
    if not m:
        return None
    token = m.group(0).replace(",", ".")  # normalize decimal comma

    try:
        return float(token)
    except ValueError:
        return None

def build_merged_lookup(ws):
    lookup = {}
    for rng in ws.merged_cells.ranges:
        tl = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                lookup[(r, c)] = tl
    return lookup

def get_value_resolving_merged(ws, r, c, merged_lookup):
    tl = merged_lookup.get((r, c))
    if tl and tl != (r, c):
        r, c = tl
    return ws.cell(row=r, column=c).value

def find_start_markers(ws, search_cols=range(1, 12)) -> List[Dict[str, Any]]:
    """Rows with a 'Lotto:' cell in A..K; copy the exact text."""
    markers = []
    for r in range(1, ws.max_row + 1):
        for c in search_cols:
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and RE_HAS_LOTTO_COLON.search(val):
                markers.append({"row": r, "coord": cell.coordinate, "text": val.strip()})
                break
    return markers

def row_rightmost_number(ws, row: int) -> Optional[float]:
    """Find the rightmost numeric value (typed or parsable) on a given row."""
    rightmost: Optional[float] = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        f = to_float(v)
        if f is not None:
            rightmost = f
    return rightmost

def find_end_row_and_number(ws, start_row: int) -> Optional[Dict[str, Any]]:
    """
    Find first row > start_row that contains the word 'Lotto' but NOT 'Lotto:'.
    Return {'row': int, 'number': float, 'text': str|None} or None.
    """
    for r in range(start_row + 1, ws.max_row + 1):
        row_has_lotto_word = False
        row_has_lotto_colon = False
        row_text_sample = None

        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                if RE_HAS_LOTTO_WORD.search(v):
                    row_has_lotto_word = True
                    row_text_sample = row_text_sample or v.strip()
                if RE_HAS_LOTTO_COLON.search(v):
                    row_has_lotto_colon = True

        if row_has_lotto_word and not row_has_lotto_colon:
            num = row_rightmost_number(ws, r)
            return {"row": r, "number": num, "text": row_text_sample}

    return None

def sum_priority_columns(ws, col_letters, r1: int, r2: int, merged_lookup) -> float:
    """
    For each row in [r1..r2]:
      - Take the first numeric among columns (priority order, e.g. ['AB','AC','AD'])
      - Ignore empty/0 values
      - Ignore numbers > 1000 (banner rows)
      - Sum the rest
    """
    if r1 > r2:
        return 0.0

    col_idxs = [column_index_from_string(cl) for cl in col_letters]
    total = 0.0

    for r in range(r1, r2 + 1):
        row_value = None
        for c in col_idxs:
            v = get_value_resolving_merged(ws, r, c, merged_lookup)
            f = to_float(v)
            if f is not None and 0 < abs(f) <= 1000:
                row_value = f
                break  # respect priority order
        if row_value is not None:
            total += row_value

    return total

def process(file: Path, sheet_name: Optional[str] = None, value_col: str = "AC"):
    wb = load_workbook(file, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    starts = find_start_markers(ws)          # rows with 'Lotto:' in A..K
    results = []

    for s in starts:
        end = find_end_row_and_number(ws, s["row"])   # row w/ 'Lotto' but not 'Lotto:'
        start_sum_row = s["row"] + 1
        end_sum_row = (end["row"] - 1) if end else ws.max_row
        merged_lookup = build_merged_lookup(ws)  # build once (do this near top of process)
        total = sum_priority_columns(ws, ["AC", "AD", "AH"], start_sum_row, end_sum_row, merged_lookup)

        results.append({
            "LottoText": s["text"],                     # exact text from the first Lotto: cell
            "SumUnderAC": total,
            "EndLottoNumber": (end["number"] if end else None),
            "StartCell": s["coord"],
            "EndRow": (end["row"] if end else None),
            "EndRowText": (end["text"] if end else None),
        })
    return results

if __name__ == "__main__":
    # edit these lines (or wire up argparse)
    xlsx = Path(r"C:\Users\ddamen\Downloads\cijfers.xlsx")
    sheet = "Table 1"   # or None to use active sheet
    value_col = "AB"

    rows = process(xlsx, sheet, value_col)

    if not rows:
        print("No blocks found.")
    else:
        print(f"\nFound {len(rows)} block(s):\n")
        print(f"{'StartCell':<10} | {'Sum':<12} | {'End #':<10} | Lotto Text")
        print("-" * 80)
        for r in rows:
            print(f"{r['StartCell']:<10} | {r['SumUnderAC']:<12.2f} | "
                  f"{(r['EndLottoNumber'] if r['EndLottoNumber'] is not None else ''):<10} | "
                  f"{r['LottoText']}")

        out = xlsx.with_name("lotto_summary.xlsx")
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Lotto Summary"

        # headers
        headers = ["StartCell", "SumUnderAC", "EndLottoNumber", "LottoText", "EndRow", "EndRowText"]
        ws_out.append(headers)

        # data rows
        for r in rows:
            ws_out.append([
                r.get("StartCell", ""),
                r.get("SumUnderAC", ""),
                r.get("EndLottoNumber", ""),
                r.get("LottoText", ""),
                r.get("EndRow", ""),
                r.get("EndRowText", "")
            ])

        # optional: make columns a bit wider
        for col in ["A","B","C","D","E","F"]:
            ws_out.column_dimensions[col].width = 25

        wb_out.save(out)
        print(f"\nSaved Excel file: {out}")
